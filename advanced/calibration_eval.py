#!/usr/bin/env python3
"""Judge MSQA model answers and compute confidence calibration summaries.

This script is the cleaned release version of the internal calibration
evaluation script. It intentionally contains no API keys, private proxy URLs, or
local absolute paths. Configure the judge endpoint through environment variables
or command-line arguments:

    set GEMINI_API_KEY=your_api_key
    set GEMINI_BASE_URL=https://generativelanguage.googleapis.com/v1beta/openai
    python judge_existing_model_calibration.py --workbook path\to\results.xlsx

Expected workbook layout:
- The first 11 sheets contain language-specific QA results.
- Row 1 contains model names above repeated model-output columns.
- Row 2 contains metadata headers plus repeated "response"/"confidence" fields.
- A sheet named "Judge SP" contains the judge prompt template.

Outputs:
- task1_gemini_judgements.jsonl: per-answer judge cache, resumable by unique key.
- task1_bucket_summary.csv: model-level accuracy by confidence bucket.
- task1_model_summary.csv: model-level accuracy, average confidence, and ECE.
- task1_language_bucket_summary.csv: language/model confidence-bucket summary.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
import warnings
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Set, Tuple

import numpy as np


# Older openpyxl releases may still touch deprecated numpy aliases.
with warnings.catch_warnings():
    warnings.simplefilter("ignore", FutureWarning)
    if not hasattr(np, "float"):
        np.float = float  # type: ignore[attr-defined]
    if not hasattr(np, "int"):
        np.int = int  # type: ignore[attr-defined]
    if not hasattr(np, "bool"):
        np.bool = bool  # type: ignore[attr-defined]

import openpyxl
import requests


LANGUAGE_SHEET_COUNT = 11
JUDGE_PROMPT_SHEET = "Judge SP"
PREFERRED_WORKBOOK_NAMES = [
    "多语言校准度测试-全结果.xlsx",
    "multilingual_calibration_full_results.xlsx",
]

DEFAULT_JUDGE_MODEL = "gemini-3.1-pro-preview"
FALLBACK_JUDGE_MODELS = [
    "gemini-3-flash-preview",
    "gemini-3-flash",
    "gemini-2.5-pro",
    "gemini-2.5-flash",
]
DEFAULT_GEMINI_BASE_URL = "https://generativelanguage.googleapis.com/v1beta/openai"

RESULTS_JSONL = "task1_gemini_judgements.jsonl"
SUMMARY_BUCKET_CSV = "task1_bucket_summary.csv"
SUMMARY_MODEL_CSV = "task1_model_summary.csv"
SUMMARY_LANGUAGE_BUCKET_CSV = "task1_language_bucket_summary.csv"


@dataclass(frozen=True)
class Example:
    sheet: str
    row_number: int
    item_id: str
    prompt_id: str
    culture_circle: str
    category: str
    prompt: str
    gold_answer: str
    question_zh: str
    answer_zh: str
    model_name: str
    model_col: int
    response: str
    confidence: int

    @property
    def unique_key(self) -> str:
        return f"{self.sheet}::{self.prompt_id}::{self.model_name}"


def normalize_base_url(base_url: str) -> str:
    cleaned = base_url.rstrip("/")
    if not cleaned.endswith("/v1") and not cleaned.endswith("/openai"):
        cleaned = f"{cleaned}/v1"
    return cleaned


class JudgeClient:
    """Direct HTTP client for OpenAI-compatible chat-completions endpoints."""

    def __init__(self, api_key: Optional[str] = None, base_url: Optional[str] = None) -> None:
        self.api_key = api_key or os.getenv("MSQA_JUDGE_API_KEY") or os.getenv("GEMINI_API_KEY")
        if not self.api_key:
            raise ValueError("Missing judge API key. Set MSQA_JUDGE_API_KEY / GEMINI_API_KEY or pass --api-key.")

        self.base_url = normalize_base_url(
            base_url
            or os.getenv("MSQA_JUDGE_BASE_URL")
            or os.getenv("GEMINI_BASE_URL", DEFAULT_GEMINI_BASE_URL)
        )
        self.session = requests.Session()
        self.session.headers.update(
            {
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
            }
        )

    def call_text_only(
        self,
        prompt: str,
        model: str,
        max_tokens: int,
        reasoning_effort: Optional[str],
    ) -> str:
        payload: Dict[str, object] = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": max_tokens,
        }
        if reasoning_effort:
            payload["extra_body"] = {"reasoning_effort": reasoning_effort}

        response = self.session.post(
            f"{self.base_url}/chat/completions",
            json=payload,
            timeout=120,
        )
        try:
            response.raise_for_status()
        except requests.HTTPError as exc:
            details = response.text[:500]
            raise RuntimeError(
                f"Judge API request failed with HTTP {response.status_code}: {details}"
            ) from exc

        data = response.json()
        try:
            return data["choices"][0]["message"]["content"] or ""
        except (KeyError, IndexError, TypeError) as exc:
            raise RuntimeError(f"Unexpected judge response payload: {data}") from exc


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Use an LLM judge to score MSQA responses and compute calibration metrics."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=find_workbook(script_dir),
        help="Path to the QA workbook. Defaults to a recognized .xlsx in this directory.",
    )
    parser.add_argument("--judge-model", default=DEFAULT_JUDGE_MODEL)
    parser.add_argument(
        "--api-key",
        default=None,
        help="Judge API key. Prefer GEMINI_API_KEY; do not hard-code keys in source files.",
    )
    parser.add_argument(
        "--base-url",
        default=None,
        help="OpenAI-compatible base URL. Prefer GEMINI_BASE_URL for local runs.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=script_dir,
        help="Directory for jsonl/csv outputs. Default: script directory.",
    )
    parser.add_argument("--limit", type=int, default=None, help="Only judge N pending examples.")
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Delete the existing JSONL cache and re-judge all examples.",
    )
    parser.add_argument("--max-tokens", type=int, default=1024)
    parser.add_argument(
        "--reasoning-effort",
        default=None,
        choices=["minimal", "medium", "high"],
        help="Optional extra_body reasoning_effort for gateways that support it.",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=0.0,
        help="Optional delay after each successful judge call.",
    )
    return parser.parse_args()


def find_workbook(base_dir: Path) -> Path:
    for workbook_name in PREFERRED_WORKBOOK_NAMES:
        preferred = base_dir / workbook_name
        if preferred.exists():
            return preferred

    candidates = sorted(
        path for path in base_dir.glob("*.xlsx") if not path.name.startswith("~$")
    )
    if not candidates:
        raise FileNotFoundError(
            f"No workbook found under {base_dir}. Pass one explicitly with --workbook."
        )
    return candidates[0]


def first_non_empty_cell(ws: "openpyxl.worksheet.worksheet.Worksheet") -> str:
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if value is not None and str(value).strip():
                return str(value).strip()
    raise ValueError(f"No non-empty cell found in sheet {ws.title!r}")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_confidence(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    try:
        numeric = int(round(float(value)))
    except (TypeError, ValueError):
        return None
    return max(0, min(100, numeric))


def get_optional_value(row: Sequence[object], header_map: Dict[str, int], header: str) -> str:
    index = header_map.get(header)
    if index is None or index >= len(row):
        return ""
    return normalize_text(row[index])


def infer_model_columns(row1: Sequence[object]) -> List[Tuple[int, str]]:
    model_columns = [
        (idx, normalize_text(value))
        for idx, value in enumerate(row1)
        if normalize_text(value)
    ]
    if not model_columns:
        raise ValueError("No model columns found in sheet header row")
    return model_columns


def infer_response_confidence_columns(
    row2: Sequence[object],
    model_col: int,
) -> Optional[Tuple[int, int]]:
    header_at_model = normalize_text(row2[model_col] if model_col < len(row2) else None)
    next_header = normalize_text(row2[model_col + 1] if model_col + 1 < len(row2) else None)
    second_next_header = normalize_text(row2[model_col + 2] if model_col + 2 < len(row2) else None)

    if header_at_model == "result" and next_header == "response" and second_next_header == "confidence":
        return model_col + 1, model_col + 2
    if header_at_model == "response" and next_header == "confidence":
        return model_col, model_col + 1
    return None


def iter_examples(workbook_path: Path) -> Iterator[Example]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    language_sheet_names = workbook.sheetnames[:LANGUAGE_SHEET_COUNT]

    for sheet_name in language_sheet_names:
        ws = workbook[sheet_name]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        header_map = {
            normalize_text(value): idx
            for idx, value in enumerate(row2)
            if normalize_text(value)
        }
        model_columns = infer_model_columns(row1)

        required_headers = ["_id", "prompt_id", "prompt", "answer"]
        missing_headers = [header for header in required_headers if header not in header_map]
        if missing_headers:
            raise KeyError(f"Sheet {sheet_name!r} missing headers: {missing_headers}")

        for row_number, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not normalize_text(row[0] if row else None):
                continue

            for model_col, model_name in model_columns:
                response_confidence_cols = infer_response_confidence_columns(row2, model_col)
                if response_confidence_cols is None:
                    continue

                response_col, confidence_col = response_confidence_cols
                response = normalize_text(row[response_col]) if response_col < len(row) else ""
                confidence = normalize_confidence(row[confidence_col] if confidence_col < len(row) else None)
                if not response or confidence is None:
                    continue

                yield Example(
                    sheet=sheet_name,
                    row_number=row_number,
                    item_id=get_optional_value(row, header_map, "_id"),
                    prompt_id=get_optional_value(row, header_map, "prompt_id"),
                    culture_circle=get_optional_value(row, header_map, "culture_circle"),
                    category=get_optional_value(row, header_map, "category"),
                    prompt=get_optional_value(row, header_map, "prompt"),
                    gold_answer=get_optional_value(row, header_map, "answer"),
                    question_zh=get_optional_value(row, header_map, "question_zh"),
                    answer_zh=get_optional_value(row, header_map, "answer_zh"),
                    model_name=model_name,
                    model_col=model_col + 1,
                    response=response,
                    confidence=confidence,
                )


def load_existing_keys(results_path: Path) -> Set[str]:
    return {str(record["unique_key"]) for record in read_results(results_path) if "unique_key" in record}


def load_existing_unique_model_counts(results_path: Path) -> Counter:
    model_keys: Dict[str, Set[str]] = {}
    for payload in read_results(results_path):
        key = payload.get("unique_key")
        model_name = payload.get("model_name")
        if key and model_name:
            model_keys.setdefault(str(model_name), set()).add(str(key))
    return Counter({model_name: len(keys) for model_name, keys in model_keys.items()})


def build_judge_prompt(judge_template: str, example: Example) -> str:
    payload = {
        "language_sheet": example.sheet,
        "prompt_id": example.prompt_id,
        "culture_circle": example.culture_circle,
        "category": example.category,
        "question": example.prompt,
        "question_zh": example.question_zh,
        "gold_target": example.gold_answer,
        "gold_target_zh": example.answer_zh,
        "predicted_answer": example.response,
    }
    return (
        f"{judge_template}\n\n"
        "Evaluate the following QA item.\n"
        "Use the scoring rule above and follow the mandatory output format exactly.\n\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )


def parse_judge_response(text: str) -> Tuple[int, str]:
    lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
    if not lines:
        raise ValueError("Judge response has no non-empty lines")

    match = re.match(r"^([01])(?:\b|$)", lines[0])
    if not match:
        raise ValueError(f"Judge response first line is not 0/1: {lines[0]!r}")

    score = int(match.group(1))
    explanation = "\n".join(lines[1:]).strip()
    return score, explanation


def judge_example(
    client: JudgeClient,
    judge_prompt: str,
    judge_model: str,
    max_tokens: int,
    reasoning_effort: Optional[str],
    retries: int = 3,
) -> Tuple[int, str, str]:
    last_error: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            raw_text = client.call_text_only(
                judge_prompt,
                model=judge_model,
                max_tokens=max_tokens,
                reasoning_effort=reasoning_effort,
            )
            score, explanation = parse_judge_response(raw_text)
            return score, explanation, raw_text
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            if "Unauthorized" in str(exc) or "HTTP 4" in str(exc):
                break
            if attempt < retries:
                time.sleep(min(4, attempt))

    raise RuntimeError(f"Judge call failed after {retries} attempt(s): {last_error}") from last_error


def should_fallback_model(error: Exception) -> bool:
    message = str(error).lower()
    fallback_markers = [
        "model_not_found",
        "model not found",
        "no available channel",
        "has no available channels",
        "empty judge response",
        "judge response has no non-empty lines",
    ]
    return any(marker in message for marker in fallback_markers)


def build_judge_model_candidates(primary_model: str) -> List[str]:
    candidates = [primary_model]
    for model in FALLBACK_JUDGE_MODELS:
        if model not in candidates:
            candidates.append(model)
    return candidates


def confidence_bucket(confidence: int) -> str:
    lower = min((confidence // 10) * 10, 90)
    upper = 100 if lower == 90 else lower + 9
    return f"{lower:02d}-{upper:02d}"


def write_jsonl_record(path: Path, payload: Dict[str, object]) -> None:
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def read_results(path: Path) -> List[Dict[str, object]]:
    if not path.exists():
        return []

    results: List[Dict[str, object]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                results.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return results


def mean(values: Sequence[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def expected_calibration_error(group: Sequence[Dict[str, object]]) -> float:
    if not group:
        return 0.0

    bucketed: Dict[str, List[Dict[str, object]]] = {}
    for record in group:
        bucketed.setdefault(str(record["confidence_bucket"]), []).append(record)

    total = len(group)
    ece = 0.0
    for records_in_bucket in bucketed.values():
        bucket_weight = len(records_in_bucket) / total
        accuracy = mean([int(item["judge_score"]) for item in records_in_bucket]) * 100.0
        avg_confidence = mean([float(item["confidence"]) for item in records_in_bucket])
        ece += bucket_weight * abs(avg_confidence - accuracy)
    return ece


def summarize_results(
    records: Sequence[Dict[str, object]],
) -> Tuple[List[Dict[str, object]], List[Dict[str, object]], List[Dict[str, object]]]:
    by_model_bucket: Dict[Tuple[str, str], List[Dict[str, object]]] = {}
    by_model: Dict[str, List[Dict[str, object]]] = {}
    by_language_model_bucket: Dict[Tuple[str, str, str], List[Dict[str, object]]] = {}

    for record in records:
        model = str(record["model_name"])
        sheet = str(record["sheet"])
        bucket = str(record["confidence_bucket"])
        by_model_bucket.setdefault((model, bucket), []).append(record)
        by_model.setdefault(model, []).append(record)
        by_language_model_bucket.setdefault((sheet, model, bucket), []).append(record)

    bucket_rows = [
        make_bucket_row(model, bucket, group)
        for (model, bucket), group in sorted(by_model_bucket.items())
    ]

    model_rows: List[Dict[str, object]] = []
    for model, group in sorted(by_model.items()):
        scores = [int(item["judge_score"]) for item in group]
        confidences = [float(item["confidence"]) for item in group]
        accuracy = mean(scores) * 100.0
        avg_confidence = mean(confidences)
        model_rows.append(
            {
                "model_name": model,
                "sample_count": len(group),
                "correct_count": sum(scores),
                "accuracy_pct": round(accuracy, 2),
                "avg_confidence_pct": round(avg_confidence, 2),
                "calibration_gap_pct": round(avg_confidence - accuracy, 2),
                "ece_pct": round(expected_calibration_error(group), 2),
            }
        )

    language_bucket_rows = [
        {"sheet": sheet, **make_bucket_row(model, bucket, group)}
        for (sheet, model, bucket), group in sorted(by_language_model_bucket.items())
    ]
    return bucket_rows, model_rows, language_bucket_rows


def make_bucket_row(
    model: str,
    bucket: str,
    group: Sequence[Dict[str, object]],
) -> Dict[str, object]:
    scores = [int(item["judge_score"]) for item in group]
    confidences = [float(item["confidence"]) for item in group]
    accuracy = mean(scores) * 100.0
    avg_confidence = mean(confidences)
    return {
        "model_name": model,
        "confidence_bucket": bucket,
        "sample_count": len(group),
        "correct_count": sum(scores),
        "accuracy_pct": round(accuracy, 2),
        "avg_confidence_pct": round(avg_confidence, 2),
        "calibration_gap_pct": round(avg_confidence - accuracy, 2),
    }


def write_csv(path: Path, rows: Sequence[Dict[str, object]], fieldnames: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_summary_outputs(output_dir: Path, records: Sequence[Dict[str, object]]) -> None:
    bucket_rows, model_rows, language_bucket_rows = summarize_results(records)
    write_csv(
        output_dir / SUMMARY_BUCKET_CSV,
        bucket_rows,
        [
            "model_name",
            "confidence_bucket",
            "sample_count",
            "correct_count",
            "accuracy_pct",
            "avg_confidence_pct",
            "calibration_gap_pct",
        ],
    )
    write_csv(
        output_dir / SUMMARY_MODEL_CSV,
        model_rows,
        [
            "model_name",
            "sample_count",
            "correct_count",
            "accuracy_pct",
            "avg_confidence_pct",
            "calibration_gap_pct",
            "ece_pct",
        ],
    )
    write_csv(
        output_dir / SUMMARY_LANGUAGE_BUCKET_CSV,
        language_bucket_rows,
        [
            "sheet",
            "model_name",
            "confidence_bucket",
            "sample_count",
            "correct_count",
            "accuracy_pct",
            "avg_confidence_pct",
            "calibration_gap_pct",
        ],
    )


def select_pending_examples(
    examples: Iterable[Example],
    existing_keys: Set[str],
    existing_model_counts: Counter,
    limit: Optional[int],
) -> Tuple[List[Example], int]:
    all_examples = list(examples)
    expected_model_counts = Counter(example.model_name for example in all_examples)
    completed_models = {
        model_name
        for model_name, expected_count in expected_model_counts.items()
        if existing_model_counts.get(model_name, 0) >= expected_count
    }

    pending: List[Example] = []
    for example in all_examples:
        if example.model_name in completed_models:
            continue
        if example.unique_key in existing_keys:
            continue
        pending.append(example)
        if limit is not None and len(pending) >= limit:
            break

    return pending, len(all_examples)


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    results_path = output_dir / RESULTS_JSONL
    if args.overwrite and results_path.exists():
        results_path.unlink()

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if JUDGE_PROMPT_SHEET not in workbook.sheetnames:
        raise KeyError(f"Workbook missing judge prompt sheet: {JUDGE_PROMPT_SHEET!r}")
    judge_template = first_non_empty_cell(workbook[JUDGE_PROMPT_SHEET])

    existing_keys = load_existing_keys(results_path)
    existing_model_counts = load_existing_unique_model_counts(results_path)
    client = JudgeClient(api_key=args.api_key, base_url=args.base_url)
    pending_examples, total_examples = select_pending_examples(
        iter_examples(workbook_path),
        existing_keys=existing_keys,
        existing_model_counts=existing_model_counts,
        limit=args.limit,
    )

    print(f"Workbook: {workbook_path}")
    print(f"Judge prompt sheet: {JUDGE_PROMPT_SHEET}")
    print(f"Judge model: {args.judge_model}")
    print(f"Base URL: {client.base_url}")
    print(f"Existing cached judgements: {len(existing_keys)}")
    print(f"Pending examples to judge this run: {len(pending_examples)}")
    print(f"Total examples in workbook: {total_examples}")

    for index, example in enumerate(pending_examples, start=1):
        prompt = build_judge_prompt(judge_template, example)
        candidate_models = build_judge_model_candidates(args.judge_model)
        last_error: Optional[RuntimeError] = None

        for candidate_index, candidate_model in enumerate(candidate_models):
            try:
                score, explanation, raw_text = judge_example(
                    client=client,
                    judge_prompt=prompt,
                    judge_model=candidate_model,
                    max_tokens=args.max_tokens,
                    reasoning_effort=args.reasoning_effort,
                )
                used_model = candidate_model
                break
            except RuntimeError as exc:
                last_error = exc
                if not should_fallback_model(exc):
                    raise
                if candidate_index < len(candidate_models) - 1:
                    print(
                        f"Judge model unavailable for {example.prompt_id}: {candidate_model}. "
                        f"Trying {candidate_models[candidate_index + 1]} next."
                    )
        else:
            raise RuntimeError(f"All judge model candidates failed: {last_error}") from last_error

        record = {
            "unique_key": example.unique_key,
            "sheet": example.sheet,
            "row_number": example.row_number,
            "item_id": example.item_id,
            "prompt_id": example.prompt_id,
            "culture_circle": example.culture_circle,
            "category": example.category,
            "model_name": example.model_name,
            "confidence": example.confidence,
            "confidence_bucket": confidence_bucket(example.confidence),
            "gold_answer": example.gold_answer,
            "predicted_answer": example.response,
            "judge_model_candidates": candidate_models,
            "judge_model": used_model,
            "judge_score": score,
            "judge_explanation": explanation,
            "judge_raw_output": raw_text,
        }
        write_jsonl_record(results_path, record)
        print(
            f"[{index}/{len(pending_examples)}] "
            f"{example.sheet} {example.prompt_id} {example.model_name} "
            f"conf={example.confidence} score={score} judge={used_model}"
        )
        if args.sleep_seconds > 0:
            time.sleep(args.sleep_seconds)

    all_records = read_results(results_path)
    write_summary_outputs(output_dir, all_records)
    print(f"\nSaved detailed judgements to: {results_path}")
    print(f"Saved bucket summary to: {output_dir / SUMMARY_BUCKET_CSV}")
    print(f"Saved model summary to: {output_dir / SUMMARY_MODEL_CSV}")
    print(f"Saved language bucket summary to: {output_dir / SUMMARY_LANGUAGE_BUCKET_CSV}")


if __name__ == "__main__":
    main()
